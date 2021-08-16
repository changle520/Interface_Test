# -*- coding: utf-8 -*- 
# @Time : 2020/11/27 11:50 
# @Author : qian 
# @File : readexceltools.py
from openpyxl import load_workbook


class Excel_read:
    def __init__(self, filepath, sheetname):
        self.wk = load_workbook(filepath)
        # self.sheet=self.wk.get_sheet_by_name(sheetname)
        self.sheet = self.wk[sheetname]  # 新方法
        self.maxrow = self.sheet.max_row  # 最大行数
        self.maxcolumn = self.sheet.max_column  # 最大列数

    def get_dict_data(self):
        if self.maxrow <= 1:
            raise ValueError('Excel文件内容总行数≤1')
        else:
            dic_list = []
            for i in range(2, self.maxrow + 1):
                d = {}
                d['rowNum'] = i
                for j in range(1, self.maxcolumn + 1):
                    key = self.sheet.cell(1, j).value
                    if self.sheet.cell(i, j).value == None:
                        d[key] = ''
                    else:
                        d[key] = self.sheet.cell(i, j).value
                dic_list.append(d)
            return (dic_list)


if __name__ == '__main__':
    import os
    datapath = os.path.join(os.path.dirname(__file__), '../test_data/ParamExcel')
    filepath = os.path.join(datapath, 'api_case.xlsx')
    print(filepath)
    #a = Excel_read(filepath, 'Sheet1')
    a = Excel_read(filepath, '测试详情')
    for i in a.get_dict_data():
        print(i)
        print(type(i))

